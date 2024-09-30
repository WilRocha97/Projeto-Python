# -*- coding: utf-8 -*-
from tkinter.filedialog import askopenfilename, askdirectory, Tk
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from functools import wraps
from pathlib import Path

from pyautogui import alert, confirm, hotkey, pixel
from threading import Thread, Lock
from io import BytesIO
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from email.message import EmailMessage
from smtplib import SMTP_SSL
from email.utils import make_msgid
import time, psutil, socket, random, os, re, traceback, tempfile, contextlib, OpenSSL.crypto, PySimpleGUI as sg, pandas as pd

try:
    from mysql.connector.errors import DatabaseError
    from _comum.mysql_comum import _create_tables_task_watch, _update_script_status, _update_historico_status
except:
    print('Biblioteca "mysql-connector-python" não instalada.')
    pass

# Crie um lock global
lock = Lock()

dados_contadores = "Dados Contadores.txt"

# variáveis globais
e_dir = Path('execução')
_headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}


def barra_de_status(func):
    sg.LOOK_AND_FEEL_TABLE['tema'] = {'BACKGROUND': '#cacaca',
                                      'TEXT': '#000000',
                                      'INPUT': '#ffffff',
                                      'TEXT_INPUT': '#ffffff',
                                      'SCROLL': '#ffffff',
                                      'BUTTON': ('#000000', '#ffffff'),
                                      'PROGRESS': ('#ffffff', '#ffffff'),
                                      'BORDER': 0,
                                      'SLIDER_DEPTH': 0,
                                      'PROGRESS_DEPTH': 0}
    
    @wraps(func)
    def wrapper():
        def image_to_data(im):
            """
            Image object to bytes object.
            : Parameters
              im - Image object
            : Return
              bytes object.
            """
            with BytesIO() as output:
                im.save(output, format="PNG")
                data = output.getvalue()
            return data
        
        def update_window_elements(window, updates):
            while True:
                try:
                    # Loop para aplicar as atualizações
                    for key, update_args in updates.items():
                        window[key].update(**update_args)
                    break
                except:
                    pass
        
        filename = 'V:/Setor Robô/Scripts Python/_comum/Assets/processando_spin_bars.gif'
        im = Image.open(filename)
        
        filename_check = 'V:/Setor Robô/Scripts Python/_comum/Assets/check.png'
        im_check = Image.open(filename_check)
        
        filename_error = 'V:/Setor Robô/Scripts Python/_comum/Assets/error.png'
        im_error = Image.open(filename_error)
        
        index = 0
        frames = im.n_frames
        im.seek(index)
        
        sg.theme('tema')  # Define o tema do PySimpleGUI
        # sg.theme_previewer()
        # Layout da janela
        layout = [[
            sg.pin(sg.Image(data=image_to_data(im), key='-Processando-', size=(38, 38), visible=False)),
            sg.pin(sg.Image(data=image_to_data(im_check), key='-Check-', size=(38, 38), visible=False)),
            sg.pin(sg.Image(data=image_to_data(im_error), key='-Error-', size=(38, 38), visible=False)),
            sg.pin(sg.Text('', key='-titulo-', visible=False, border_width=0, font=("Helvetica", 11, 'bold'))),
            sg.pin(sg.Text('', key='-Mensagens-', visible=False, font=("Helvetica", 11, 'bold'))),
            sg.pin(sg.Button('INICIAR', key='-iniciar-', border_width=0, button_color=('black', '#BFBFBF'), font=("Helvetica", 11, 'bold'))),
            sg.pin(sg.Button('FECHAR', key='-fechar-', border_width=0, button_color=('black', '#BFBFBF'), font=("Helvetica", 11, 'bold'), visible=False))
        ]]
        
        # guarda a janela na variável para manipula-la
        screen_width, screen_height = sg.Window.get_screen_size()
        window = sg.Window('', layout, transparent_color='#cacaca', no_titlebar=True, keep_on_top=True, element_justification='center', margins=(0, 0), finalize=True, size=(screen_width, 38), location=((screen_width // 2) - (screen_width // 2), 0))
        
        def run_script_thread():
            try:
                # Chama a função que executa o script
                hotkey('win', 'm')
                func(window)
            except Exception as e:
                # Obtém a pilha de chamadas de volta como uma string
                traceback_str = traceback.format_exc()
                
                # habilita e desabilita os botões conforme necessário
                updates = {'-titulo-': {'value': 'Erro :('},
                           '-Mensagens-': {'value': '', 'visible': False},
                           '-Processando-': {'visible': False},
                           '-fechar-': {'visible': True},
                           '-Error-': {'visible': True}}
                update_window_elements(window, updates)
                
                alert(f'Traceback: {traceback_str}\n\n'
                      f'Erro: {e}')
                print(f'Traceback: {traceback_str}\n\n'
                      f'Erro: {e}')
                return
            
            # habilita e desabilita os botões conforme necessário
            updates = {'-titulo-': {'value': 'Execução finalizada :)'},
                       '-Mensagens-': {'value': '', 'visible': False},
                       '-Processando-': {'visible': False},
                       '-fechar-': {'visible': True},
                       '-Check-': {'visible': True}}
            update_window_elements(window, updates)
        
        processando = 'não'
        while True:
            # captura o evento e os valores armazenados na interface
            event, values = window.read(timeout=5)
            
            r, g, b = pixel(int(screen_width / 2), 10)
            if r >= 127.5 and g >= 127.5 and b >= 127.5:
                window['-titulo-'].update(text_color='black')
                window['-Mensagens-'].update(text_color='black')
            else:
                window['-titulo-'].update(text_color='white')
                window['-Mensagens-'].update(text_color='white')
            
            if event == sg.WIN_CLOSED:
                break
            elif event == '-fechar-':
                break
            
            elif event == '-iniciar-':
                # habilita e desabilita os botões conforme necessário
                updates = {'-titulo-': {'value': 'Rotina automática.', 'visible': True},
                           '-Processando-': {'visible': True},
                           '-Mensagens-': {'visible': True},
                           '-iniciar-': {'visible': False}}
                update_window_elements(window, updates)
                
                # Cria uma nova thread para executar o script
                Thread(target=run_script_thread).start()
                processando = 'sim'
            
            elif processando == 'sim':
                index = (index + 1) % frames
                im.seek(index)
                window['-Processando-'].update(data=image_to_data(im), size=(80, 38))
        
        window.close()
    
    return wrapper


_barra_de_status = barra_de_status


def _time_execution_monitor(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        controle_historico = 'V:\\Setor Robô\\Scripts Python\\_comum\\_Monitoramento\\Histórico\\histórico.txt'
        controle_rotinas = args[0]
        nome_rotina = controle_rotinas.split('\\')[-1].replace('.txt', '')
        comeco = datetime.now()
        
        print(f"🕐 Execução iniciada as: {comeco}\n")
        with open(controle_rotinas, 'w', encoding='latin-1') as f:
            f.write('[Iniciando notifica]')
        f = open(controle_historico, 'a', encoding='latin-1')
        data_hora = datetime.now().strftime('%d/%m/%Y - %H:%M')
        f.write(f'[{data_hora}] [{nome_rotina}] [Rotina iniciada]\n')
        f.close()
        
        try:
            func(*args, **kwargs)
        except:
            # Obtém a pilha de chamadas de volta como uma string
            traceback_str = traceback.format_exc()
            print(traceback_str)
            if controle_rotinas:
                with open(controle_rotinas, 'w', encoding='latin-1') as f:
                    f.write('[Erro detectado notifica]')
                f = open(controle_historico, 'a', encoding='latin-1')
                data_hora = datetime.now().strftime('%d/%m/%Y - %H:%M')
                f.write(f'[{data_hora}] [{nome_rotina}] [Erro detectado]\n')
                f.close()
                while True:
                    time.sleep(1)
        
        print(f"\n🕞 Tempo de execução: {datetime.now() - comeco}\n🕖 Encerrado as: {datetime.now()} ")
        with open(controle_rotinas, 'w', encoding='latin-1') as f:
            f.write('[Rotina finalizada]')
        f = open(controle_historico, 'a', encoding='latin-1')
        data_hora = datetime.now().strftime('%d/%m/%Y - %H:%M')
        f.write(f'[{data_hora}] [{nome_rotina}] [Rotina finalizada]\n')
        f.close()
    
    return wrapper


def _time_execution_monitor_db(func):
    """ Decorador para monitorar os andamentos da rotina """
    
    # conecta no banco de dados
    try:
        _create_tables_task_watch()
        sem_bd = False
    except DatabaseError:
        traceback_str = traceback.format_exc()
        print(traceback_str)
        print("Erro ao conectar ao banco de dados do TaskWatch, verifique se o mesmo está em execução ou se esta maquina tem permissão para acessao.")
        sem_bd = True
        pass
    
    @wraps(func)
    def wrapper(*args, **kwargs):
        # recebe os dados da rotina para adicionar no banco de dados e poder monitora-la
        controle_rotinas = args[0]
        nome_rotina = controle_rotinas.split('\\')[-1].replace('.txt', '')
        comeco = datetime.now()
        
        print(f"🕐 Execução iniciada as: {comeco}\n")
        
        if not sem_bd:
            # se conseguir se conectar no banco atualiza a tabela de execuções e o histórico
            _update_script_status(script_name=nome_rotina, status='status-executando', andamentos='Iniciando...')
            _update_historico_status(script_name=nome_rotina, andamentos='Rotina iniciada')
        try:
            # executa o script
            func(*args, **kwargs)
        except:
            if not sem_bd:
                # se conseguir se conectar no banco atualiza a tabela de execuções e o histórico
                _update_script_status(script_name=nome_rotina, status='status-error', andamentos='Erro detectado', porcentagem='✖')
                _update_historico_status(script_name=nome_rotina, andamentos='Erro detectado')
            # Obtém a pilha de chamadas de volta como uma string
            traceback_str = traceback.format_exc()
            print(traceback_str)
            while True:
                time.sleep(1)
        
        if not sem_bd:
            # se conseguir se conectar no banco atualiza a tabela de execuções e o histórico
            _update_script_status(script_name=nome_rotina, status='status-final', andamentos='Rotina finalizada', porcentagem='✔')
            _update_historico_status(script_name=nome_rotina, andamentos='Rotina finalizada')
        print(f"\n🕞 Tempo de execução: {datetime.now() - comeco}\n🕖 Encerrado as: {datetime.now()} ")
    
    return wrapper


def _atualiza_monitor(controle, mensagem='', novo_conteudo=False, atualiza_arquivo=False):
    try:
        # atualiza o arquivo de controle para evitar que ative o alerta de ociosidade
        if atualiza_arquivo:
            with open(controle, 'r', encoding='latin-1') as f:
                conteudo_anterior = f.read()
            with open(controle, 'w', encoding='latin-1') as f:
                f.write(conteudo_anterior)
            return
        
        if not novo_conteudo:
            # atualiza o arquivo de controle com algúm aviso caso a rotina esteja mais lente do que normal devido a alguma exceção que a rotina consegue tratar sozinha
            with open(controle, 'r', encoding='latin-1') as f:
                conteudo_anterior = f.read()
                conteudos = re.compile(r'\[ (.+ Restantes) \] ').search(conteudo_anterior)
                novo_conteudo = f'[ {conteudos.group(1)} ] [{mensagem}]'
            with open(controle, 'w', encoding='latin-1') as f:
                f.write(novo_conteudo)
        else:
            with open(controle, 'w', encoding='latin-1') as f:
                f.write(f'{mensagem}')
    except AttributeError:
        with open(controle, 'w', encoding='latin-1') as f:
            f.write(f'[{mensagem}]')


def time_execution(func):
    """Decorator que mede o tempo de execução de uma função decorada com ela"""
    
    @wraps(func)
    def wrapper():
        comeco = datetime.now()
        print(f"🕐 Execução iniciada as: {comeco}\n")
        func()
        print(f"\n🕞 Tempo de execução: {datetime.now() - comeco}\n🕖 Encerrado as: {datetime.now()} ")
    
    return wrapper


_time_execution = time_execution

'''
Decorator para cenários onde pode ou não já existir uma instancia de BeatifulSoup
Cria uma instancia de BeautifulSoup, referente ao response.content 'content'
recebido, e passa para a função 'func' recebida ou simplesmente passa a
instancia de BeautifulSoup 'soup' recebida

Na declaração de uma funcao genérica utiliza-se a seguinte forma

@content_or_soup
def funcao_generica(soup):
    # faça alguma_coisa
    # return alguma_coisa ou não

Dessa forma qualquer funcao decorada com esse decorator pode receber tanto um
response.content ou uma instancia de BeautifulSoup da seguinte forma

para response.content -> funcao_generica(content=response.content)
para instancia BeautifulSoup -> funcao_generica(soup=BeautifulSoup(content, 'parser'))
'''


def content_or_soup(func):
    def wrapper(*dump, content=None, soup=None):
        if content:
            soup = BeautifulSoup(content, 'html.parser')
        return func(soup, *dump)
    
    return wrapper


_content_or_soup = content_or_soup


@contextlib.contextmanager
def pfx_to_pem(pfx_path, pfx_psw):
    # Decrypts the .pfx file to be used with requests.
    with tempfile.NamedTemporaryFile(suffix='.pem', delete=False) as t_pem:
        pfx = open(pfx_path, 'rb').read()  # --> type(pfx) > bytes
        p12 = OpenSSL.crypto.load_pkcs12(pfx, pfx_psw.encode('utf8'))
        
        if p12.get_certificate().has_expired():
            print('Certificado possivelmente vencido')
            dti = p12.get_certificate().get_notBefore()
            dtv = p12.get_certificate().get_notAfter()
            print(f'Inicio: {dti[6:8]}/{dti[4:6]}/{dti[:4]}')
            print(f'Vencimento: {dtv[6:8]}/{dtv[4:6]}/{dtv[:4]}')
        
        f_pem = open(t_pem.name, 'wb')
        f_pem.write(OpenSSL.crypto.dump_privatekey(OpenSSL.crypto.FILETYPE_PEM, p12.get_privatekey()))
        f_pem.write(OpenSSL.crypto.dump_certificate(OpenSSL.crypto.FILETYPE_PEM, p12.get_certificate()))
        ca = p12.get_ca_certificates()
        if ca is not None:
            for cert in ca:
                f_pem.write(OpenSSL.crypto.dump_certificate(OpenSSL.crypto.FILETYPE_PEM, cert))
        f_pem.close()
        
        yield t_pem.name


_pfx_to_pem = pfx_to_pem


def which_cert(cnpj, psw=None):
    """Recebe um cnpj e procura na pasta _cert por um certificado correspondente, caso 'senha is None' tenta os 8 primeiros dígitos do cnpj como senha
    Retorna o caminho do cert em caso de sucesso retorna mensagem de erro caso falha"""
    
    c_dir = os.path.join('..', '..', '_cert')
    
    pfxs = tuple(i for i in os.listdir(c_dir) if i[-4:] == '.pfx')
    
    for pfx in pfxs:
        if not psw:
            psw = re.search(r'(\d{8})', pfx)
            if psw:
                psw = psw.group(1)
            else:
                raise Exception('Falta senha no nome do cert', pfx)
        
        try:
            pfx = os.path.join(c_dir, pfx)
            with open(pfx, 'rb') as cert:
                p12 = OpenSSL.crypto.load_pkcs12(
                    cert.read(), psw.encode('utf8')
                )
        except OpenSSL.crypto.Error:
            continue
        
        cert_id = p12._friendlyname.decode()
        if cnpj in cert_id:
            return pfx, psw
    
    return 'cnpj e/ou senha não correspondentes'


_which_cert = which_cert


def atualiza_contadores():
    # obtêm a lista de usuário e senha de cada contador para a planilha dedados de algumas consultas
    f = open(dados_contadores, 'r', encoding='utf-8')
    contadores = f.readlines()
    
    contadores_dict = {}
    for contador in contadores:
        contador = contador.split('/')
        contadores_dict[contador[0]] = (contador[1], contador[2], contador[3])
    
    return contadores_dict


def _escreve_relatorio_xlsx(texto, local='execução', nome='Relatório', encode='latin-1'):
    os.makedirs(local, exist_ok=True)
    caminho_arquivo = os.path.join(local, f"{nome}.xlsx")
    
    # Adquira o lock antes de acessar o arquivo
    with lock:
        while True:
            try:
                workbook = load_workbook(caminho_arquivo)
                break
            except FileNotFoundError:
                # Se o arquivo não existir, crie-o com pandas e mantenha a formatação
                df_status = pd.DataFrame([texto])
                try:
                    df_status.to_excel(caminho_arquivo, index=False)
                except:
                    df_status.to_excel(caminho_arquivo.replace('.xlsx', ' - auxiliar.xlsx'), index=False)
                
                return
            except:
                pass
        
        sheet = workbook.active
        
        # Encontre o próximo índice de linha disponível
        nova_linha_index = sheet.max_row + 1
        
        # Crie um DataFrame a partir dos dados fornecidos
        df_novos_dados = pd.DataFrame([texto])
        
        # Adicione novas linhas à planilha
        for r_idx, row in enumerate(dataframe_to_rows(df_novos_dados, index=False, header=False), nova_linha_index):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Salve o workbook
        workbook.save(caminho_arquivo)


def escreve_relatorio_csv(texto, nome='resumo', local=e_dir, end='\n', encode='latin-1'):
    """Recebe um texto 'texto' junta com 'end' e escreve num arquivo 'nome'"""
    
    os.makedirs(local, exist_ok=True)
    
    try:
        f = open(os.path.join(local, f"{nome}.csv"), 'a', encoding=encode)
    except:
        f = open(os.path.join(local, f"{nome} - auxiliar.csv"), 'a', encoding=encode)
    
    f.write(texto + end)
    f.close()


_escreve_relatorio_csv = escreve_relatorio_csv


def escreve_header_csv(texto, nome='resumo', local=e_dir, encode='latin-1'):
    """Recebe um cabeçalho 'texto' e escreve no comeco do arquivo 'nome'"""
    
    os.makedirs(local, exist_ok=True)
    
    try:
        with open(os.path.join(local, f"{nome}.csv"), 'r', encoding=encode) as f:
            conteudo = f.read()
        with open(os.path.join(local, f"{nome}.csv"), 'w', encoding=encode) as f:
            f.write(texto + '\n' + conteudo)
    except:
        with open(os.path.join(local, f"{nome} - auxiliar.csv"), 'r', encoding=encode) as f:
            conteudo = f.read()
        with open(os.path.join(local, f"{nome} - auxiliar.csv"), 'w', encoding=encode) as f:
            f.write(texto + '\n' + conteudo)


_escreve_header_csv = escreve_header_csv


def ask_for_file(title='Abrir arquivo', filetypes='*', initialdir=os.getcwd()):
    """wrapper para askopenfilename"""
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    
    file = askopenfilename(
        title=title,
        filetypes=filetypes,
        initialdir=initialdir
    )
    
    return file if file else False


_ask_for_file = ask_for_file


def ask_for_dir(title='Abrir pasta'):
    """wrapper para askdirectory"""
    
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    
    folder = askdirectory(
        title=title,
    )
    
    return folder if folder else False


_ask_for_dir = ask_for_dir


def where_to_start(idents, encode='latin-1', file=False):
    """# Procura pelo indice do primeiro campo da última linha, que deve ser um identificador unico como cpf/cnpj/inscricao, do arquivo de resumo escolhido e o
    retorna 0 caso escolha não na caixa de texto ou não encontre o identificador na variável 'idents'
    retorna None caso caixa de texto seja fechada ou o arquivo resumo, não seja escolhido ou não consiga abrir o arquivo escolhido"""
    
    title = 'Execucao anterior'
    text = 'Deseja continuar execucao anterior?'
    
    if not file:
        res = confirm(title=title, text=text, buttons=('sim', 'não'))
        if not res:
            return None
        if res == 'não':
            return 0
        
        ftypes = [('Plain text files', '*.txt *.csv')]
        file = ask_for_file(filetypes=ftypes)
        if not file:
            return None
    
    try:
        with open(file, 'r', encoding=encode) as f:
            dados = f.readlines()
    except Exception as e:
        alert(title='Mensagem erro', text=f'Não pode abrir arquivo\n{str(e)}')
        return None
    
    try:
        elem = dados[-1].split(';')[0]
        return idents.index(elem) + 1
    except ValueError:
        return 0


_where_to_start = where_to_start


def _configura_dados(pasta_final_, andamentos, continuar_rotina=False, nova_planilha=True, planilha_dados=False, empresas_20000=False, colunas_usadas=None, colunas_filtro=None, palavras_filtro=None, filtrar_celulas_em_branco=None):
    def where_to_start_pandas(pasta_final_anterior, planilha_andamentos, df_empresas):
        print(pasta_final_anterior, planilha_andamentos)
        try:
            df_andamentos = pd.read_excel(os.path.join(pasta_final_anterior, planilha_andamentos))
        except:
            print('❗❗ Não foi encontrada nenhuma planilha de andamentos na pasta de execução anterior.\n\n'
                  f'Começando a execução a partir do primeiro índice da planilha de dados selecionada.')
            return 0
        
        # pega o valor da última linha da primeira coluna para buscar o index na planilha de dados
        ultima_linha_processada = df_andamentos.iloc[-1, 0]
        
        # Procurar esse valor na primeira coluna do segundo DataFrame
        index = df_empresas[df_empresas.iloc[:, 0] == ultima_linha_processada].index
        
        try:
            if int(index[0]) == df_empresas.last_valid_index():
                print(f'INDICE CAPTURADO: 0')
                return 0
            
            print(f'INDICE CAPTURADO: {int(index[0])}')
            print(f'ULTIMO INDICE DA PLANILHA DE DADOS: {df_empresas.last_valid_index()}')
        except IndexError:
            print(f'INDICE CAPTURADO: 0')
            return 0
        
        # Se última linha processada não for encontrada, iniciar do começo
        if not index.empty:
            return int(index[0]) + 1
        else:
            return 0
    
    def get_newer_file(folder_path, filename, latest_time, latest_file, latest_folder=None):
        # Verifica se o arquivo tem extensão .xlsx ou .xls
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(folder_path, filename)
            # Obtém a data de modificação do arquivo
            file_time = os.path.getmtime(file_path)
            # Compara com a data do arquivo mais recente encontrado até agora
            if latest_time is None or file_time > latest_time:
                latest_time = file_time
                latest_file = filename
                latest_folder = folder_path
        
        return latest_folder, latest_file, latest_time
    
    def get_latest_excel_file(folder_path='V:\\Setor Robô\\Relatórios\\ae'):
        # Inicializa variáveis para armazenar o nome do arquivo mais recente e sua data de modificação
        latest_time = None
        latest_file = None
        
        # Itera sobre todos os arquivos na pasta
        for filename in os.listdir(folder_path):
            latest_folder, latest_file, latest_time = get_newer_file(folder_path, filename, latest_time, latest_file)
        
        return os.path.join(folder_path, latest_file)
    
    def get_last_rotine_excel(folder_path):
        latest_time = None
        latest_file = None
        latest_folder = None
        
        for pasta_atual, subpastas, arquivos in os.walk(folder_path):
            # Agora você pode processar os arquivos na pasta atual normalmente
            for filename in arquivos:
                latest_folder, latest_file, latest_time = get_newer_file(pasta_atual, filename, latest_time, latest_file, latest_folder)
        
        return latest_folder, latest_file
    
    comp = datetime.now().strftime('%m-%Y')
    ano = datetime.now().strftime('%Y')
    pasta_final_ = os.path.join(pasta_final_, andamentos, ano, comp)
    contador = 0
    index = 0
    pasta_final_anterior = False
    
    if empresas_20000:
        com_20000 = '(Com empresas acima do código 20000)'
    if not empresas_20000:
        com_20000 = ''
    
    if planilha_dados == 'Não se aplica':
        index = None
        df_empresas = None
        total_empresas = None
        
        # iteração para determinar se precisa criar uma pasta nova para armazenar os resultados
        # toda vês que o programa começar as consultas uma nova pasta será criada para não sobrepor ou misturar as execuções
        while True:
            try:
                pasta_final = os.path.join(pasta_final_, f'Execução')
                os.makedirs(pasta_final)
                break
            except:
                try:
                    contador += 1
                    pasta_final = os.path.join(pasta_final_, f'Execução ({str(contador)})')
                    os.makedirs(pasta_final)
                    break
                except:
                    pass
        
        return pasta_final, index, df_empresas, total_empresas
    if not planilha_dados:
        planilha_dados = get_latest_excel_file()
    
    # verifica se existe uma planilha de dados:
    ultima_pasta, ultima_planilha = get_last_rotine_excel(pasta_final_)
    
    # se não tiver uma planilha de andamentos anterior no mês, cria uma nova
    if not ultima_planilha or not continuar_rotina:
        print('Não tem andamentos anteriores')
        while True:
            try:
                pasta_final = os.path.join(pasta_final_, f'Execução{com_20000}')
                os.makedirs(pasta_final)
                break
            except:
                try:
                    contador += 1
                    pasta_final = os.path.join(pasta_final_, f'Execução{com_20000} ({str(contador)})')
                    os.makedirs(pasta_final)
                    break
                except:
                    pass
    
    # se tiver, confere se ela é tem mais de 24 horas da última edição
    else:
        print('Tem andamentos anteriores')
        print(ultima_pasta, ultima_planilha)
        planilha_antiga = confere_data_arquivo(os.path.join(ultima_pasta, ultima_planilha))
        
        # se a planilha tiver mais de 24 horas, cria uma nova pasta final
        if planilha_antiga:
            print('Andamentos anteriores é muito antigo')
            # iteração para determinar se precisa criar uma pasta nova para armazenar os resultados
            # toda vês que o programa começar as consultas uma nova pasta será criada para não sobrepor ou misturar as execuções
            while True:
                try:
                    pasta_final = os.path.join(pasta_final_, f'Execução{com_20000}')
                    os.makedirs(pasta_final)
                    break
                except:
                    try:
                        contador += 1
                        pasta_final = os.path.join(pasta_final_, f'Execução{com_20000} ({str(contador)})')
                        os.makedirs(pasta_final)
                        break
                    except:
                        pass
        
        # se a planilha de andamentos anterior não tiver mais de 24 horas da ultima edição, utiliza ela para capturar o último indice e continuar a execução
        else:
            print('Serão usados os andamentos anteriores')
            pasta_final_anterior = ultima_pasta
            pasta_final = ultima_pasta
    
    # abrir a planilha de dados
    df_empresas, situacao = open_dados(nova_planilha, andamentos, empresas_20000, pasta_final, planilha_dados, colunas_usadas, colunas_filtro, palavras_filtro, filtrar_celulas_em_branco)
    
    if not situacao:
        return False, False, False, False
    
    total_empresas = int(df_empresas.shape[0])
    
    # se na verificação anterior da planilha de andamentos for encontrada uma planilha e ela não for mais antiga que 24 horas, busca o último indice dela para continuar a rotina atual
    if pasta_final_anterior:
        print('Continuando andamentos anteriores, pegando ultimo indice')
        planilha_andamentos = f'{andamentos}.xlsx'
        # obtêm o índice do último andamento da execução anterior para continuar
        index = where_to_start_pandas(pasta_final_anterior, planilha_andamentos, df_empresas)
        print(index)
        total_empresas = total_empresas - index
    
    return pasta_final, index, df_empresas, total_empresas


def open_dados(nova_planilha, andamentos, empresas_20000, pasta_final, planilha_dados, colunas_usadas, colunas_filtro, palavras_filtro, filtrar_celulas_em_branco):
    dados_final = os.path.join(pasta_final, 'Dados.xlsx')
    encode = 'latin-1'
    
    # modelo de lista com as colunas que serão usadas na rotina
    # colunas_usadas = ['column1', 'column2', 'column3']
    
    if nova_planilha:
        print('pasta: ', planilha_dados)
        df = pd.read_excel(planilha_dados)
        
        # coluna com os códigos do ae
        coluna_codigo = 'Codigo'
        
        # filtra as colunas
        try:
            if not empresas_20000:
                # cria um novo df apenas com empresas a baixo do código 20.000
                df_filtrada = df[df[coluna_codigo] <= 20000]
            else:
                # cria um novo df apenas com empresas a cima do código 20.000
                df_filtrada = df[df[coluna_codigo] >= 20000]
            
            # filtra as células de colunas específicas que contenham palavras especificas
            if palavras_filtro and colunas_filtro:
                for count, coluna_para_filtrar in enumerate(colunas_filtro):
                    df_filtrada = df_filtrada[df_filtrada[coluna_para_filtrar].str.contains(palavras_filtro[count], case=False, na=False)]
            
            df_filtrada = df_filtrada[colunas_usadas]
        except KeyError:
            print(f'❗❗ Erro ao buscar as colunas na planilha base selecionada: {planilha_dados}\n\n'
                  f'Verifique se a planilha contem as colunas necessárias para a execução da rotina')
            return False, False
        
        if filtrar_celulas_em_branco:
            df_filtrada = df_filtrada.dropna(subset=filtrar_celulas_em_branco)
            # df_filtrada = df_filtrada.fillna('vazio')
        else:
            # remove linha com células vazias
            df_filtrada = df_filtrada.dropna(axis=0, how='any')
        
        # Converte a coluna 'CNPJ' para string e remova a parte decimal '.0'. Preencha com zeros à esquerda para garantir 14 dígitos
        df_filtrada['CNPJ'] = df_filtrada['CNPJ'].astype(str).str.replace(r'\.0', '', regex=True).str.zfill(14)
        
        if andamentos == 'Consulta Débitos Estaduais' or andamentos == 'Consulta CND Não Inscritos':
            contadores_dict = atualiza_contadores()
            # Substituir valores com base no dicionário apenas se o valor estiver presente no dicionário
            df_filtrada['Perfil'] = 'vazio'
            
            # Função para atualizar os valores das colunas com base no dicionário de mapeamento
            def atualizar_valores(row):
                if row['PostoFiscalContador'] in contadores_dict:
                    return contadores_dict[row['PostoFiscalContador']]
                else:
                    return row['PostoFiscalUsuario'], row['PostoFiscalSenha'], 'contribuinte'
            
            # Função para atualizar os valores das colunas com base no dicionário de mapeamento
            def substitui_contribuinte(row):
                if row['Perfil'] == 'contribuinte':
                    return contadores_dict['principal']
                else:
                    return row['PostoFiscalUsuario'], row['PostoFiscalSenha'], row['Perfil']
            
            # Aplicar a função para atualizar os valores das colunas
            df_filtrada[['PostoFiscalUsuario', 'PostoFiscalSenha', 'Perfil']] = df_filtrada.apply(atualizar_valores, axis=1, result_type='expand')
            
            # 5. Deletar a coluna 'contador'
            df_filtrada.drop(columns=['PostoFiscalContador'], inplace=True)
            
            # 3. Deletar linhas com células vazias na coluna 'senha'
            df_filtrada = df_filtrada.dropna(subset=['PostoFiscalSenha'])
            
            # Aplicar a função para atualizar os valores das colunas
            df_filtrada[['PostoFiscalUsuario', 'PostoFiscalSenha', 'Perfil']] = df_filtrada.apply(substitui_contribuinte, axis=1, result_type='expand')
            
            # Ordene o DataFrame com base na coluna desejada
            df_filtrada = df_filtrada.sort_values(by=['PostoFiscalUsuario', 'CNPJ'], ascending=[True, True])
            
            # remove linha com células vazias
            df_filtrada = df_filtrada.dropna(axis=0, how='any')
            
            # Remover linhas que contenham alguma palavra específica na coluna 'PostoFiscalUsuario'
            for texto in ['ISENTO', 'BAIXADO']:
                df_filtrada = df_filtrada[~df_filtrada['PostoFiscalUsuario'].str.contains(texto, case=False, na=False)]
        
        if andamentos == 'Consulta Débitos Municipais Jundiaí' or andamentos == 'Consulta Débitos Municipais Valinhos' or andamentos == 'Consulta Débitos Municipais Vinhedo':
            # Remover linhas que contenham alguma palavra específica na coluna 'InsMunicipal'
            for texto in ['ENCERRADA', 'NÃO POSSUI', 'EM ANDAMENTO', 'ISENTO']:
                df_filtrada = df_filtrada[~df_filtrada['InsMunicipal'].str.contains(texto, case=False, na=False)]
        
        if df_filtrada.empty:
            print(f'❗❗ Não foi encontrada nenhuma empresa com os critérios necessários para essa consulta na planilha selecionada: {planilha_dados}')
            return False, False
        
        for coluna in df_filtrada.columns:
            # Remova aspas duplas
            df_filtrada[coluna] = df_filtrada[coluna].astype(str).str.replace('"', '')
            
            # Remova quebras de linha (`\n` e `\r`)
            df_filtrada[coluna] = df_filtrada[coluna].astype(str).str.replace('\n', '').str.replace('\r', '').str.replace('_x000D_', '')
        
        df_filtrada.to_excel(dados_final, index=False)
        empresas = pd.read_excel(dados_final)
    else:
        empresas = pd.read_excel(planilha_dados)
    
    print(empresas)
    return empresas, True


def open_lista_dados(i_dir='ignore', encode='latin-1', file=False, retorna_planilha=False):
    """Abre uma janela de seleção de arquivos e abre o arquivo selecionado retorna List de Tuple das linhas dividas por ';' do arquivo caso sucesso
    Retorna None caso nenhum selecionado ou erro ao ler arquivo"""
    
    ftypes = [('Plain text files', '*.txt *.csv')]
    
    if not file:
        file = ask_for_file(filetypes=ftypes, initialdir=i_dir)
        if not file:
            return False
    
    try:
        with open(file, 'r', encoding=encode) as f:
            dados = f.readlines()
    except Exception as e:
        alert(title='Mensagem erro', text=f'Não pode abrir arquivo\n{i_dir}\n{str(e)}')
        return False
    
    print('>>> usando dados de ' + file.split('/')[-1])
    if retorna_planilha:
        return list(map(lambda x: tuple(x.replace('\n', '').split(';')), dados)), file
    else:
        return list(map(lambda x: tuple(x.replace('\n', '').split(';')), dados))


_open_lista_dados = open_lista_dados


def download_file(name, response, pasta=str(e_dir / 'docs')):
    """Recebe o 'response' de um request e salva o conteudo num arquivo 'name' na pasta 'pasta' retorna True em caso de sucesso
    Levanta uma exceção em caso de erro"""
    
    pasta = str(pasta).replace('\\', '/')
    os.makedirs(pasta, exist_ok=True)
    
    with open(os.path.join(pasta, name), 'wb') as arq:
        for i in response.iter_content(100000):
            arq.write(i)


_download_file = download_file


def indice(count, total_empresas, empresa='', index=0, window=False, tempos=False, tempo_execucao=None, controle=None, usando_bd=False, nome_rotina='Rotina', mensagen_extra=''):
    """ Função para mostrar qual empresa está sendo utilizada no andamento atual.
        Calcula o tempo de execução para finalizar a rotina """
    try:
        total_empresas = len(total_empresas)
    except:
        pass
    
    previsao_termino_texto = ''
    tempo_estimado_texto = ''
    tempos_de_execucao = '[][]'
    tempo_estimado = 0
    tempo_medio_texto = ''
    dados_rotina = ''
    
    # Cria um indice para saber qual linha dos dados está
    indice_dados = f'[ {mensagen_extra}{str(count + index - 1)} de {str(total_empresas + index)} | {str((total_empresas + index) - (count + index - 1))} Restantes ]'
    
    # se tiver habilitado para calcular o tempo de execução
    if tempos:
        tempo_inicial = datetime.now()
        
        tempos.append(tempo_inicial)
        tempo_execucao_atual = int(tempos[1].timestamp()) - int(tempos[0].timestamp())
        tempos.pop(0)
        
        # verifica se o lista 'tempo_execucao' tem mais de 100 itens, se tiver, tira o primeiro para ficar somente os 100 mais recentes
        if len(tempo_execucao) > 50:
            del (tempo_execucao[0])
        
        tempo_execucao.append(tempo_execucao_atual)
        for t in tempo_execucao:
            tempo_estimado = tempo_estimado + t
        tempo_estimado = int(tempo_estimado) / int(len(tempo_execucao))
        
        # pega o tempo médio de cada execução em texto
        if tempo_estimado > 0:
            # Converter o tempo total para um objeto timedelta
            tempo_medio = timedelta(seconds=tempo_estimado)
            dias_texto, horas_texto, minutos_texto, segundos_texto = converte_tempo_em_texto(tempo_medio)
            tempo_medio_texto = f"[ Tempo médio por ciclo: {dias_texto}{horas_texto}{minutos_texto}{segundos_texto} ]"
        
        tempo_total_segundos = int((total_empresas + index) - (count + index) + 1) * int(tempo_estimado)
        tempo_estimado = tempo_execucao
        # Converter o tempo total para um objeto timedelta
        tempo_total = timedelta(seconds=tempo_total_segundos)
        
        # Extrair dias, horas e minutos do timedelta
        dias, dias_texto, horas, horas_texto, minutos, minutos_texto, segundos, segundos_texto = converte_tempo_em_texto(tempo_total, retorna_texto_numero=True)
        
        if dias > 0 or horas > 0 or minutos > 0:
            previsao_termino = tempo_inicial + tempo_total
            # Retorna o tempo no formato "dias:horas:minutos:segundos"
            previsao_termino_texto = f"[ Previsão de termino: {previsao_termino.strftime('%d/%m/%Y as %H:%M')} ]"
            tempo_estimado_texto = f"[ Tempo estimado: {dias_texto}{horas_texto}{minutos_texto} ]"
            tempos_de_execucao = f'{tempo_estimado_texto}{previsao_termino_texto}'
    
    if window:
        # se tiver a barra de status, tenta atualiza-la
        while True:
            status_2 = f'{str((count + index) - 1)} de {str(total_empresas + index)} | {str((total_empresas + index) - (count + index) + 1)} Restantes{tempos_de_execucao}'
            try:
                window['-Mensagens-'].update(status_2)
                break
            except:
                try:
                    window['-Mensagens-'].update('Buguei...')
                except:
                    pass
                print('>>> Erro ao atualizar a interface, tentando novamente...')
                pass
    
    porcentagem = ((count + index) / (total_empresas + index)) * 100
    dados_rotina = indice_dados + tempo_medio_texto + tempos_de_execucao + f'[ {int(porcentagem)}% ]'
    empresa = str(empresa).replace("('", '[ ').replace("')", ' ]').replace("',)", " ]").replace(',)', ' ]').replace("', '", ' - ')
    
    print(f'\n\n{dados_rotina}\n{empresa}')
    
    # se for utilizar algúm metodo de monitoramento dos andamentos
    if controle:
        # se usar banco de dados, atualiza os andamentos nele
        if usando_bd:
            try:
                _update_script_status(script_name=nome_rotina.replace('[', '').replace(']', '').strip(),
                                      status='status-executando',
                                      andamentos=indice_dados.replace('[', '').replace(']', '').strip(),
                                      tempo_medio=tempo_medio_texto.replace('[', '').replace(']', '').strip(),
                                      tempo_estimado=tempo_estimado_texto.replace('[', '').replace(']', '').strip(),
                                      previsao_termino=previsao_termino_texto.replace('[', '').replace(']', '').strip(),
                                      porcentagem=f'%{int(porcentagem)}')
            except DatabaseError:
                print("Erro ao conectar ao banco de dados do TaskWatch, verifique se o mesmo está em execução.")
                pass
            except:
                traceback_str = traceback.format_exc()
                print(traceback_str)
                pass
        
        # se não usar banco de dados, atualiza o arquivo txt de controle
        else:
            if dados_rotina != '':
                with open(controle, 'w', encoding='latin-1') as f:
                    f.write(dados_rotina)
    
    return tempos, tempo_estimado


_indice = indice


def converte_tempo_em_texto(tempo_total, returna_numeros=False, retorna_texto_numero=False):
    # Extrair dias, horas e minutos do timedelta
    dias = tempo_total.days
    horas = tempo_total.seconds // 3600
    minutos = (tempo_total.seconds % 3600) // 60
    segundos = tempo_total.seconds % 60
    
    dias_texto = ''
    horas_texto = ''
    minutos_texto = ''
    segundos_texto = ''
    
    if dias == 1:
        dias_texto = f'{dias} dia '
    elif dias > 1:
        dias_texto = f'{dias} dias '
    if horas == 1:
        horas_texto = f'{horas} hora '
    elif horas > 1:
        horas_texto = f'{horas} horas '
    if minutos == 1:
        minutos_texto = f'{minutos} minuto '
    elif minutos > 1:
        minutos_texto = f'{minutos} minutos '
    if segundos == 1:
        segundos_texto = f'{segundos} segundo'
    elif segundos > 1:
        segundos_texto = f'{segundos} segundos'
    
    if returna_numeros:
        return dias, horas, minutos, segundos
    if retorna_texto_numero:
        return dias, dias_texto, horas, horas_texto, minutos, minutos_texto, segundos, segundos_texto
    else:
        return dias_texto, horas_texto, minutos_texto, segundos_texto


def remove_emojis(text):
    emoji_pattern = re.compile("["
                               u"\U0001F600-\U0001F64F"  # emoticons
                               u"\U0001F300-\U0001F5FF"  # símbolos e pictogramas
                               u"\U0001F680-\U0001F6FF"  # transportes e símbolos de mapa
                               u"\U0001F1E0-\U0001F1FF"  # bandeiras de país
                               "]+", flags=re.UNICODE)
    text = emoji_pattern.sub(r'', text)
    text = text[0:-1]
    return text


_remove_emojis = remove_emojis


def remove_espacos(text):
    string = re.compile(r"\s\s")
    text = string.sub(r'', text)
    text = text[0:-1]
    return text


_remove_espacos = remove_espacos


def generate_random_number(lista_controle):
    """ Gera um número aleatório de 10 dígitos que não contem na lista"""
    while True:
        controle = str(random.randint(10 ** 9, 10 ** 10 - 1))
        if controle not in lista_controle:
            lista_controle.append(controle)
            return controle, lista_controle


def concatena(variavel, quantidade, posicao, caractere):
    variavel = str(variavel)
    if posicao == 'depois':
        # concatena depois
        while len(str(variavel)) < quantidade:
            variavel += str(caractere)
    if posicao == 'antes':
        # concatena antes
        while len(str(variavel)) < quantidade:
            variavel = str(caractere) + str(variavel)
    
    return variavel


_concatena = concatena


def _kill_process_by_name(process_name):
    """Encerra todos os processos com o nome especificado."""
    for proc in psutil.process_iter(['name']):
        try:
            if proc.info['name'].lower() == process_name.lower():
                proc.kill()
                print(f"Processo '{process_name}' encerrado.")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass


def _divide_list(lst, n, total_empresas):
    if total_empresas < 20:
        n = 1
    
    # Divide a lista em n partes
    k, m = divmod(len(lst), n)
    return [lst[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n)], n


def _get_host_name():
    # Get the hostname of the machine
    return str(socket.gethostname()).upper()


def _envia_email(destinatario, andamentos, pasta_final, anexo, cc_list=False):
    email = "roboautorobot@gmail.com"
    senha_do_email = 'ymgwnihtsjcipfhv'
    
    msg = EmailMessage()  #
    msg['Subject'] = f'Robô: {andamentos}'
    msg['From'] = email
    msg['To'] = destinatario
    if cc_list:
        msg['Cc'] = cc_list  # Adiciona destinatários em cópia
    
    cid = make_msgid()[1:-1]  # Gera um Content-ID único para a imagem
    corpo_html = f'''<html>
                        <body>
                            <p>Olá, tenha um excelente dia.</p>
                            <p>Realizamos a execução do processo automatizado <b>{andamentos}</b></p>
                            <p>Os resultados estão no seguinte diretório: <b>{pasta_final}</b><br>
                            Copie e cole esse caminho na barra de endereço do explorador de arquivos ou na barra de pesquisa do windows.</p>
                            <p>Caso houver alguma particularidade na execução peço que nos responda este e-mail com o detalhamento dela.</p>
                            <p>Todo processo automatizado é gerado a partir dos dados enviados, por se tratar de uma rotina automatizada não há intervenção humana além do tratamento dos dados. Por este motivo é necessária à conferência dos dados gerados pelo robô por parte do solicitante.</p>
                            <p>Em caso de dúvida, continuo à sua disposição.</p>
                            <p>Atenciosamente,</p>
                            <img src="cid:{cid}" alt="Assinatura" />
                        </body>
                    </html>'''
    
    msg.set_content(corpo_html, subtype='html')
    
    # Adicionando a imagem embutida
    with open('V:\\Setor Robô\\Scripts Python\\_comum\\Assets\\assinatura.png', 'rb') as img:
        img_data = img.read()
        msg.add_related(img_data, maintype='image', subtype='png', cid=cid)
    
    if anexo:
        with open(anexo, 'rb') as anexo_file:
            content = anexo_file.read()
            filename = anexo.split('/')[-1]
            msg.add_attachment(content, maintype='application', subtype='octet-stream', filename=filename)
    
    with SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email, senha_do_email)
        smtp.send_message(msg)


def confere_data_arquivo(file_path):
    """Confere se o arquivo foi criado ou modificado no dia anterior"""
    
    # Obtém o timestamp de criação do arquivo
    timestamp_criacao = os.path.getmtime(file_path)
    
    # Converte o timestamp em um objeto datetime
    data_criacao = datetime.fromtimestamp(timestamp_criacao)
    
    # Obtém a data e hora atuais
    agora = datetime.now()
    
    # Calcula a diferença entre a data atual e a data de criação
    diferenca = agora - data_criacao
    
    # Verifica se a diferença é maior que 24 horas
    planilha_velha = diferenca > timedelta(hours=24)
    
    if planilha_velha:
        return True
    else:
        return False