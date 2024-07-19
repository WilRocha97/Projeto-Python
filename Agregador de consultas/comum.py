# -*- coding: utf-8 -*-
import random, chromedriver_autoinstaller, shutil, re, os, pandas as pd, matplotlib.pyplot as plt
from tkinter.filedialog import askdirectory, Tk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import SessionNotCreatedException, WebDriverException
from anticaptchaofficial.imagecaptcha import *
from anticaptchaofficial.recaptchav2proxyless import *
from anticaptchaofficial.hcaptchaproxyless import *
from PIL import Image
from io import BytesIO

dados_anticaptcha = "Dados/Dados AC.txt"
dados_contadores = "Dados/Contadores.txt"

# Gera um número aleatório de 8 dígitos
id_execucao = random.randint(10000000, 99999999)

arquivo_log = f'Log_{id_execucao}'
controle_rotinas = f'Log/Controle_{id_execucao}.txt'
controle_botoes = f'Log/Buttons_{id_execucao}.txt'
dados_elementos = f'Log/window_values_{id_execucao}.json'


def ask_for_dir(title='Salvar onde?'):
    root = Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    
    folder = askdirectory(
        title=title,
    )
    
    return folder if folder else False


def update_window_elements(window, updates):
    # Loop para aplicar as atualizações
    for key, update_args in updates.items():
        while True:
            try:
                window[key].update(**update_args)
                break
            except:
                if key == '-progressbar-':
                    try:
                        window[key].update_bar(update_args['value'], max=update_args['max'])
                        break
                    except:
                        pass
                else:
                    pass


def concatena(variavel, quantidade, posicao, caractere):
    # função para concatenar strings colocando caracteres no começo ou no final
    variavel = str(variavel)
    if posicao == 'depois':
        # concatena depois
        while len(str(variavel)) < quantidade:
            variavel += str(caractere)
    if posicao == 'antes':
        # concatena antes
        while len(str(variavel)) < quantidade:
            variavel = str(caractere) + str(variavel)
    
    return variavel


def configura_navegador(window_principal, pasta=None, retorna_options=False, timeout=90, desabilita_seguranca=False):
    # opções para fazer com que o chrome trabalhe em segundo plano (opcional)
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--window-size=1920,1080')
    # options.add_argument(f'--proxy-server=http://{proxy_ip_port}')
    # options.add_argument("--start-maximized")
    
    if desabilita_seguranca:
        options.add_argument("--ignore-certificate-errors")
        options.add_argument("--allow-running-insecure-content")
        options.add_argument("--disable-web-security")
        
    # se a execução for salvar alguma coisa direto do navegador, configura a pasta padrão
    if pasta:
        options.add_experimental_option('prefs', {
            "download.default_directory": pasta.replace('/', '\\'),  # Change default directory for downloads
            "download.prompt_for_download": False,  # To auto download the file
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,  # It will not show PDF directly in chrome
            "profile.default_content_setting_values.automatic_downloads": 1  # download multiple files
        })
    
    # se for apenas retornar as opções do navegador para executar em outra função personalizada
    if retorna_options:
        return options
    
    update_window_elements(window_principal, {'-Mensagens_2-': {'value': 'Iniciando ambiente da consulta, aguarde...'}})
    return initialize_chrome(timeout, options)


def initialize_chrome(timeout, options=webdriver.ChromeOptions()):
    service = Service('Chrome driver')
    while True:
        for pasta_atual, subpastas, arquivos in os.walk('Chrome driver'):
            # Agora você pode processar os arquivos na pasta atual normalmente
            for file in arquivos:
                caminho_completo = os.path.join(pasta_atual, file)
                service = Service(caminho_completo)
                print(caminho_completo)
            
        if not options:
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
        
        options.add_argument("--ignore-certificate-errors")
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        # retorna o chromedriver aberto
        try:
            print('>>> Inicializando Chromedriver...')
            driver = webdriver.Chrome(options=options, service=service)
            driver.set_page_load_timeout(timeout)
            break
        except SessionNotCreatedException:
            print('>>> Atualizando Chromedriver...')
            shutil.rmtree('Chrome driver')
            time.sleep(1)
            os.makedirs('Chrome driver', exist_ok=True)
            # biblioteca para baixar o chromedriver atualizado
            chromedriver_autoinstaller.install(path='Chrome driver')
        except WebDriverException:
            print('>>> Baixando Chromedriver...')
            os.makedirs('Chrome driver', exist_ok=True)
            # biblioteca para baixar o chromedriver atualizado
            chromedriver_autoinstaller.install(path='Chrome driver')
            
    return True, driver


def find_by_id(item, driver):
    try:
        driver.find_element(by=By.ID, value=item)
        return True
    except:
        return False


def find_by_path(item, driver):
    try:
        driver.find_element(by=By.XPATH, value=item)
        return True
    except:
        return False


def send_input(elem_id, data, driver):
    while True:
        try:
            elem = driver.find_element(by=By.ID, value=elem_id)
            elem.send_keys(data)
            break
        except:
            pass


def send_input_xpath(elem_path, data, driver):
    while True:
        try:
            elem = driver.find_element(by=By.XPATH, value=elem_path)
            elem.send_keys(data)
            break
        except:
            pass


def get_info_post(soup):
    # captura infos para realizar um post
    soup = BeautifulSoup(soup, 'html.parser')
    infos = [
        soup.find('input', attrs={'id': '__VIEWSTATE'}),
        soup.find('input', attrs={'id': '__VIEWSTATEGENERATOR'}),
        soup.find('input', attrs={'id': '__EVENTVALIDATION'})
    ]
    
    # state, generator, validation
    return tuple(info.get('value', '') for info in infos if info)


def new_session_fazenda_driver(window_principal, user, pwd, perfil, retorna_driver=False, options='padrão', timeout=90):
    return cookies, sid


def solve_recaptcha(window_principal, data):
    return g_response
   

def solve_text_captcha(window_principal, driver, captcha_element, element_type='id'):
    return captcha_text


def indice(count, empresa, total_empresas, index, window_principal, tempos, tempo_execucao):
    data_atual = datetime.now()
    tempo_estimado_texto = ''
    
    # print(count, 'primeiro')
    tempo_estimado = 0
    if type(total_empresas) == list:
        quantidade_total_empresas = len(total_empresas)
    else:
        quantidade_total_empresas = int(total_empresas)
    
    # captura a hora atual e coloca em uma lista para calcular o tempo de execução do andamento atual e depois deleta o primeiro item da lista 'tempos', pois só pode haver 2
    tempos.append(data_atual)
    tempo_execucao_atual = int(tempos[1].timestamp()) - int(tempos[0].timestamp())
    tempos.pop(0)
    
    # verifica se o lista 'tempo_execucao' tem mais de 100 itens, se tiver, tira o primeiro para ficar somente os 100 mais recentes
    if len(tempo_execucao) > 100:
        del(tempo_execucao[0])
    
    # adiciona o tempo de execução atual na lista com os tempos anteriores para calcular a média de tempo de execução dos andamentos
    tempo_execucao.append(tempo_execucao_atual)
    for t in tempo_execucao:
        tempo_estimado = tempo_estimado + t
    tempo_estimado = int(tempo_estimado) / int(len(tempo_execucao))
    
    # multiplica o tempo médio de execução dos andamentos pelo número de andamentos que faltam executar para obter o tempo estimado em segundos
    tempo_total_segundos = int(quantidade_total_empresas - (count + index) + 1) * float(tempo_estimado)
    # Converter o tempo total para um objeto timedelta
    tempo_total = timedelta(seconds=tempo_total_segundos)
    
    # Extrair dias, horas e minutos do timedelta
    dias = tempo_total.days
    horas = tempo_total.seconds // 3600
    minutos = (tempo_total.seconds % 3600) // 60
    segundos = (tempo_total.seconds % 3600) % 60
    
    # Retorna o tempo no formato "dias:horas:minutos:segundos"
    dias_texto = ''
    horas_texto = ''
    minutos_texto = ''
    segundos_texto = ''
    
    if dias == 1: dias_texto = f' {dias} dia'
    elif dias > 1: dias_texto = f' {dias} dias'
    if horas == 1: horas_texto = f' {horas} hora'
    elif horas > 1: horas_texto = f' {horas} horas'
    if minutos == 1: minutos_texto = f' {minutos} minuto'
    elif minutos > 1: minutos_texto = f' {minutos} minutos'
    
    if dias > 0 or horas > 0 or minutos > 0:
        previsao_termino = data_atual + tempo_total
        # Retorna o tempo no formato "dias:horas:minutos:segundos"
        tempo_estimado_texto = f"  |  Tempo estimado:{dias_texto}{horas_texto}{minutos_texto}{segundos_texto}  |  Previsão de termino: {previsao_termino.strftime('%d/%m/%Y as %H:%M')}"
    
    print(f'\n\n[{empresa}]')
    update_window_elements(window_principal, {'-progressbar-': {'visible': True}})
    
    # print(count, index, 'segundo')
    updates = {'-Mensagens-': {'value': f'{str((count + index) - 1)} de {str(quantidade_total_empresas)}  |  {str(quantidade_total_empresas - (count + index) + 1)} Restantes{tempo_estimado_texto}'},
               '-progressbar-': {'value': (count + index) - 1, 'max': int(quantidade_total_empresas)},
               '-Progresso_texto-': {'value': str(round(float((count + index) - 1) / (int(quantidade_total_empresas)) * 100, 1)) + '%'}}
    update_window_elements(window_principal, updates)

    print(f'{str((count + index) - 1)} de {str(quantidade_total_empresas)}  |  {str(quantidade_total_empresas - (count + index) + 1)} Restantes{tempo_estimado_texto}')
    
    tempo_estimado = tempo_execucao
    return tempos, tempo_estimado


def open_dados(window_principal, situacao_dados, andamentos, empresas_20000, pasta_final, planilha_dados, colunas_usadas, colunas_filtro, palavras_filtro, filtrar_celulas_em_branco):
    dados_final = os.path.join(pasta_final, 'Dados.xlsx')
    encode = 'latin-1'
    
    # modelo de lista com as colunas que serão usadas na rotina
    # colunas_usadas = ['column1', 'column2', 'column3']
    
    if situacao_dados == '-nova_planilha-':
        df = pd.read_excel(planilha_dados)
        
        # coluna com os códigos do ae
        coluna_codigo = 'Codigo'
        
        # filtra as colunas
        try:
            if empresas_20000 == ' - (Empresas com o código menor que 20.000)':
                # cria um novo df apenas com empresas a baixo do código 20.000
                df_filtrada = df[df[coluna_codigo] <= 20000]
            elif empresas_20000 == ' - (Empresas com o código maior que 20.000)':
                # cria um novo df apenas com empresas a cima do código 20.000
                df_filtrada = df[df[coluna_codigo] >= 20000]
            else:
                df_filtrada = df
            
            # filtra as células de colunas específicas que contenham palavras especificas
            if palavras_filtro and colunas_filtro:
                for count, coluna_para_filtrar in enumerate(colunas_filtro):
                    df_filtrada = df_filtrada[df_filtrada[coluna_para_filtrar].str.contains(palavras_filtro[count], case=False, na=False)]
            
            df_filtrada = df_filtrada[colunas_usadas]
        except KeyError:
            window_principal.write_event_value('-ERRO_CRIAR_PLANILHA-', 'alerta')
            return False, False
        
        if filtrar_celulas_em_branco:
            df_filtrada = df_filtrada.dropna(subset=filtrar_celulas_em_branco)
            # df_filtrada = df_filtrada.fillna('vazio')
        else:
            # remove linha com células vazias
            df_filtrada = df_filtrada.dropna(axis=0, how='any')
        
        # Converte a coluna 'CNPJ' para string e remova a parte decimal '.0'. Preencha com zeros à esquerda para garantir 14 dígitos
        df_filtrada['CNPJ'] = df_filtrada['CNPJ'].astype(str).str.replace(r'\.0', '', regex=True).str.zfill(14)
        
        if andamentos == 'Consulta Débitos Estaduais - Situação do Contribuinte' or andamentos == 'Consulta Certidão Negativa de Débitos Tributários Não Inscritos':
            contadores_dict = atualiza_contadores()
            # Substituir valores com base no dicionário apenas se o valor estiver presente no dicionário
            df_filtrada['Perfil'] = 'vazio'
            
            # Função para atualizar os valores das colunas com base no dicionário de mapeamento
            def atualizar_valores(row):
                if row['PostoFiscalContador'] in contadores_dict:
                    return contadores_dict[row['PostoFiscalContador']]
                else:
                    return row['PostoFiscalUsuario'], row['PostoFiscalSenha'], 'contribuinte'
            
            # Aplicar a função para atualizar os valores das colunas
            df_filtrada[['PostoFiscalUsuario', 'PostoFiscalSenha', 'Perfil']] = df_filtrada.apply(atualizar_valores, axis=1, result_type='expand')
            
            # 5. Deletar a coluna 'contador'
            df_filtrada.drop(columns=['PostoFiscalContador'], inplace=True)
            
            # 3. Deletar linhas com células vazias na coluna 'senha'
            df_filtrada = df_filtrada.dropna(subset=['PostoFiscalSenha'])
            
            # Ordene o DataFrame com base na coluna desejada
            df_filtrada = df_filtrada.sort_values(by=['Perfil', 'PostoFiscalUsuario', 'CNPJ'], ascending=[True, True, True])
            
            # remove linha com células vazias
            df_filtrada = df_filtrada.dropna(axis=0, how='any')
            
            # Remover linhas que contenham alguma palavra específica na coluna 'PostoFiscalUsuario'
            for texto in ['ISENTO', 'BAIXADO']:
                df_filtrada = df_filtrada[~df_filtrada['PostoFiscalUsuario'].str.contains(texto, case=False, na=False)]
                
        if andamentos == 'Consulta Débitos Municipais Jundiaí' or andamentos == 'Consulta Débitos Municipais Valinhos' or andamentos == 'Consulta Débitos Municipais Vinhedo':
            # Remover linhas que contenham alguma palavra específica na coluna 'InsMunicipal'
            for texto in ['ENCERRADA', 'NÃO POSSUI', 'EM ANDAMENTO', 'ISENTO']:
                df_filtrada = df_filtrada[~df_filtrada['InsMunicipal'].str.contains(texto, case=False, na=False)]
            
        if df_filtrada.empty:
            window_principal.write_event_value('-PLANILHA_VAZIA-', 'alerta')
            return False, False
        
        for coluna in df_filtrada.columns:
            # Remova aspas duplas
            df_filtrada[coluna] = df_filtrada[coluna].str.replace('"', '')
            
            # Remova quebras de linha (`\n` e `\r`)
            df_filtrada[coluna] = df_filtrada[coluna].str.replace('\n', '').str.replace('\r', '').str.replace('_x000D_', '')
        
        df_filtrada.to_excel(dados_final, index=False)
        empresas = pd.read_excel(dados_final)
    else:
        empresas = pd.read_excel(planilha_dados)
    
    print(empresas)
    return empresas, True


def escreve_doc(texto, local='Log', nome=arquivo_log, encode='latin-1'):
    os.makedirs(local, exist_ok=True)
    try:
        os.remove(os.path.join(local, f"{nome}.txt"))
    except:
        pass
    try:
        os.remove(os.path.join(local, f"{nome} - auxiliar.txt"))
    except:
        pass
    
    try:
        f = open(os.path.join(local, f"{nome}.txt"), 'a', encoding=encode)
    except:
        f = open(os.path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding=encode)
    try:
        f.write(str(texto))
        f.close()
    except UnicodeEncodeError:
        try:
            f = open(os.path.join(local, f"{nome}.txt"), 'a', encoding='utf-8')
        except:
            f = open(os.path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding='utf-8')
        f.write(str(texto))
        f.close()


def escreve_relatorio_xlsx(texto, local, nome='Relatório', encode='latin-1'):
    os.makedirs(local, exist_ok=True)
    caminho_arquivo = os.path.join(local, f"{nome}.xlsx")
    
    try:
        # Abra o workbook existente
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active
        
        # Encontre o próximo índice de linha disponível
        nova_linha_index = sheet.max_row + 1
        
        # Crie um DataFrame a partir dos dados fornecidos
        df_novos_dados = pd.DataFrame([texto])
        
        # Adicione novas linhas à planilha
        for r_idx, row in enumerate(dataframe_to_rows(df_novos_dados, index=False, header=False), nova_linha_index):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Salve o workbook
        workbook.save(caminho_arquivo)
    
    except FileNotFoundError:
        # Se o arquivo não existir, crie-o com pandas e mantenha a formatação
        df_status = pd.DataFrame([texto])
        try:
            df_status.to_excel(caminho_arquivo, index=False)
        except:
            df_status.to_excel(caminho_arquivo.replace('.xlsx', ' - auxiliar.xlsx'), index=False)
    

def configura_dados(window_principal, continuar_rotina, empresas_20000, situacao_dados, planilha_dados, pasta_final_, andamentos, colunas_usadas=None, colunas_filtro=None, palavras_filtro=None, filtrar_celulas_em_branco=None):
    def where_to_start(pasta_final_anterior, planilha_andamentos, df_empresas):
        if not os.path.isdir(pasta_final_anterior):
            return 0
        
        try:
            df_andamentos = pd.read_excel(os.path.join(pasta_final_anterior, planilha_andamentos))
        except:
            window_principal.write_event_value('-PLANILHA_ANTERIOR_NAO_EXISTE-', 'alerta')
            return 0
        
        # pega o valor da última linha da primeira coluna para buscar o index na planilha de dados
        ultima_linha_processada = df_andamentos.iloc[-1, 0]
        
        # Procurar esse valor na primeira coluna do segundo DataFrame
        index = df_empresas[df_empresas.iloc[:, 0] == ultima_linha_processada].index
        
        try:
            if int(index[0]) == df_empresas.last_valid_index():
                print(f'INDICE CAPTURADO: 0')
                return 0
            
            print(f'INDICE CAPTURADO: {int(index[0])}')
            print(f'ULTIMO INDICE DA PLANILHA DE DADOS: {df_empresas.last_valid_index()}')
        except IndexError:
            print(f'INDICE CAPTURADO: 0')
            return 0
            
        # Se última linha processada não for encontrada, iniciar do começo
        if not index.empty:
            return int(index[0]) + 1
        else:
            return 0
    
    comp = datetime.now().strftime('%m-%Y')
    pasta_final_ = os.path.join(pasta_final_, andamentos, comp)
    contador = 0
       
    # iteração para determinar se precisa criar uma pasta nova para armazenar os resultados
    # toda vês que o programa começar as consultas uma nova pasta será criada para não sobrepor ou misturar as execuções
    while True:
        cr = open(controle_rotinas, 'r', encoding='utf-8').read()
        if cr == 'ENCERRAR':
            return False, False, False, False
        try:
            pasta_final = os.path.join(pasta_final_, f'Execução{empresas_20000}')
            os.makedirs(pasta_final)
            pasta_final_anterior = False
            break
        except:
            try:
                contador += 1
                pasta_final = os.path.join(pasta_final_, f'Execução{empresas_20000} ({str(contador)})')
                os.makedirs(pasta_final)
                if contador - 1 < 1:
                    pasta = f'Execução{empresas_20000}'
                else:
                    pasta = f'Execução{empresas_20000} ({str(contador - 1)})'
                pasta_final_anterior = os.path.join(pasta_final_, pasta)
                break
            except:
                pass

    if planilha_dados != 'Não se aplica':
        # abrir a planilha de dados
        update_window_elements(window_principal, {'-Mensagens_2-': {'value': 'Criando dados para a consulta...'}})
        df_empresas, situacao = open_dados(window_principal, situacao_dados, andamentos, empresas_20000, pasta_final, planilha_dados, colunas_usadas, colunas_filtro, palavras_filtro, filtrar_celulas_em_branco)
        
        if not situacao:
            return False, False, False, False
         
        if pasta_final_anterior:
            if continuar_rotina == '-continuar_rotina-':
                planilha_andamentos = f'{andamentos}.xlsx'
                # obtêm o índice do último andamento da execução anterior para continuar
                index = where_to_start(pasta_final_anterior, planilha_andamentos, df_empresas)
                print(index)
            else:
                index = 0
        else:
            index = 0
            
        total_empresas = int(df_empresas.shape[0])
    else:
        index = None
        df_empresas = None
        total_empresas = None
    
    return pasta_final, index, df_empresas, total_empresas


def atualiza_contadores():
    # obtêm a lista de usuário e senha de cada contador para a planilha dedados de algumas consultas
    f = open(dados_contadores, 'r', encoding='utf-8')
    contadores = f.readlines()
    
    contadores_dict = {}
    for contador in contadores:
        contador = contador.split('/')
        contadores_dict[contador[0]] = (contador[1], contador[2], contador[3])
    
    return contadores_dict


def download_file(name, response, pasta):
    # função para salvar um arquivo retornado de uma requisição
    pasta = str(pasta).replace('\\', '/')
    os.makedirs(pasta, exist_ok=True)
    
    with open(os.path.join(pasta, name), 'wb') as arq:
        for i in response.iter_content(100000):
            arq.write(i)
