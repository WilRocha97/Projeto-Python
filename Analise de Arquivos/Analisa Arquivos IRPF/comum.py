# -*- coding: utf-8 -*-
import os, pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
from datetime import datetime

controle_rotinas = 'Log/Controle.txt'


def concatena(variavel, quantidade, posicao, caractere):
    # função para concatenar strings colocando caracteres no começo ou no final
    variavel = str(variavel)
    if posicao == 'depois':
        # concatena depois
        while len(str(variavel)) < quantidade:
            variavel += str(caractere)
    if posicao == 'antes':
        # concatena antes
        while len(str(variavel)) < quantidade:
            variavel = str(caractere) + str(variavel)
    
    return variavel


def indice(count, empresa, total_empresas, index, window_principal, tempos, tempo_execucao):
    data_atual = datetime.now()
    tempo_estimado_texto = ''
    
    # print(count, 'primeiro')
    tempo_estimado = 0
    if type(total_empresas) == list:
        quantidade_total_empresas = len(total_empresas)
    else:
        quantidade_total_empresas = int(total_empresas)
    
    # captura a hora atual e coloca em uma lista para calcular o tempo de execução do andamento atual e depois deleta o primeiro item da lista 'tempos', pois só pode haver 2
    tempos.append(data_atual)
    tempo_execucao_atual = int(tempos[1].timestamp()) - int(tempos[0].timestamp())
    tempos.pop(0)
    
    # verifica se o lista 'tempo_execucao' tem mais de 100 itens, se tiver, tira o primeiro para ficar somente os 100 mais recentes
    if len(tempo_execucao) > 100:
        del(tempo_execucao[0])
    
    # adiciona o tempo de execução atual na lista com os tempos anteriores para calcular a média de tempo de execução dos andamentos
    tempo_execucao.append(tempo_execucao_atual)
    for t in tempo_execucao:
        tempo_estimado = tempo_estimado + t
    tempo_estimado = int(tempo_estimado) / int(len(tempo_execucao))
    
    # multiplica o tempo médio de execução dos andamentos pelo número de andamentos que faltam executar para obter o tempo estimado em segundos
    tempo_total_segundos = int(quantidade_total_empresas - (count + index) + 1) * float(tempo_estimado)
    # Converter o tempo total para um objeto timedelta
    tempo_total = timedelta(seconds=tempo_total_segundos)
    
    # Extrair dias, horas e minutos do timedelta
    dias = tempo_total.days
    horas = tempo_total.seconds // 3600
    minutos = (tempo_total.seconds % 3600) // 60
    segundos = (tempo_total.seconds % 3600) % 60
    
    # Retorna o tempo no formato "dias:horas:minutos:segundos"
    dias_texto = ''
    horas_texto = ''
    minutos_texto = ''
    segundos_texto = ''
    
    if dias == 1: dias_texto = f' {dias} dia'
    elif dias > 1: dias_texto = f' {dias} dias'
    if horas == 1: horas_texto = f' {horas} hora'
    elif horas > 1: horas_texto = f' {horas} horas'
    if minutos == 1: minutos_texto = f' {minutos} minuto'
    elif minutos > 1: minutos_texto = f' {minutos} minutos'
    
    if dias > 0 or horas > 0 or minutos > 0:
        previsao_termino = data_atual + tempo_total
        # Retorna o tempo no formato "dias:horas:minutos:segundos"
        tempo_estimado_texto = f"  |  Tempo estimado:{dias_texto}{horas_texto}{minutos_texto}{segundos_texto}  |  Previsão de termino: {previsao_termino.strftime('%d/%m/%Y as %H:%M')}"
    
    print(f'\n\n[{empresa}]')
    # print(count, index, 'segundo')
    window_principal['-progressbar-'].update(visible=True)
    window_principal['-Mensagens-'].update(f'{str((count + index) - 1)} de {str(quantidade_total_empresas)}  |  {str(quantidade_total_empresas - (count + index) + 1)} Restantes{tempo_estimado_texto}')
    window_principal['-progressbar-'].update_bar((count + index) - 1, max=int(quantidade_total_empresas))
    window_principal['-Progresso_texto-'].update(str(round(float((count + index) - 1) / (int(quantidade_total_empresas)) * 100, 1)) + '%')
    window_principal.refresh()
    print(f'{str((count + index) - 1)} de {str(quantidade_total_empresas)}  |  {str(quantidade_total_empresas - (count + index) + 1)} Restantes{tempo_estimado_texto}')
    
    tempo_estimado = tempo_execucao
    return tempos, tempo_estimado


def escreve_doc(texto, local='Log', nome='Log', encode='latin-1'):
    os.makedirs(local, exist_ok=True)
    try:
        os.remove(os.path.join(local, 'Log.txt'))
    except:
        pass
    
    try:
        f = open(os.path.join(local, f"{nome}.txt"), 'a', encoding=encode)
    except:
        f = open(os.path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding=encode)
    try:
        f.write(str(texto))
        f.close()
    except UnicodeEncodeError:
        try:
            f = open(os.path.join(local, f"{nome}.txt"), 'a', encoding='utf-8')
        except:
            f = open(os.path.join(local, f"{nome} - auxiliar.txt"), 'a', encoding='utf-8')
        f.write(str(texto))
        f.close()


def escreve_relatorio_xlsx(texto, local, nome='Relatório', encode='latin-1'):
    os.makedirs(local, exist_ok=True)
    caminho_arquivo = os.path.join(local, f"{nome}.xlsx")
    
    try:
        # Abra o workbook existente
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active
        
        # Encontre o próximo índice de linha disponível
        nova_linha_index = sheet.max_row + 1
        
        # Crie um DataFrame a partir dos dados fornecidos
        df_novos_dados = pd.DataFrame([texto])
        
        # Adicione novas linhas à planilha
        for r_idx, row in enumerate(dataframe_to_rows(df_novos_dados, index=False, header=False), nova_linha_index):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Salve o workbook
        workbook.save(caminho_arquivo)
    
    except FileNotFoundError:
        # Se o arquivo não existir, crie-o com pandas e mantenha a formatação
        df_status = pd.DataFrame([texto])
        try:
            df_status.to_excel(caminho_arquivo, index=False)
        except:
            df_status.to_excel(caminho_arquivo.replace('.xlsx', ' - auxiliar.xlsx'), index=False)
